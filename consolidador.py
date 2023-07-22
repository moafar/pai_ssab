'''
Este script recorre la carpeta donde se encuentra ubicado el archivo .py
y procesa los archivos .xlsm que encuentra creando un solo dataframe unificado.
El resultado lo guarda en el libro Excel "consolidado.xlsx" que debe existir 
en el mismo directorio donde se encuentra el script

''' 

import os
import pandas as pd
import datetime
import openpyxl

nombre_hoja = "Registro Diario"

# Obtener la ruta del directorio actual (Get Current Working Directory)
# o sea, el directorio de este archivo .py
directorio_actual = os.getcwd()

# Obtener una lista de los archivos y directorios en el directorio actual
lista_archivos = os.listdir(directorio_actual)

# Crear el DataFrame global del municipio, con 257 columnas
df_bga = pd.DataFrame(columns=range(257))

# Recorrer la lista de archivos y directorios
for archivo in lista_archivos:
    ruta_archivo = os.path.join(directorio_actual, archivo)
    if os.path.isfile(ruta_archivo) and archivo.endswith(".xlsm"):
        print(" ********************************************")
        print(datetime.datetime.now())
        nombreIPS = archivo.split()[0]
        print("Nombre IPS:", nombreIPS)
        print("Archivo:", archivo)
        df = pd.read_excel(archivo, sheet_name=nombre_hoja, skiprows=2, engine='openpyxl', header=None)
        
        # Selecciono únicamente las filas diligenciadas, verificando si la columna 1 (fecha de atención) está diligenciada
        df2 = df[df.iloc[:, 1].notnull()]

        # Elimino la última fila (cargada con "fin" en cada columna)
        df3 = df2.drop(df.index[-1])

        # Agrego la columna para identificar la IPS
        df3[len(df3.columns)] = nombreIPS

        print("Datos cargados.  El DataFrame de "+nombreIPS+" tiene", df3.shape[0], "filas")

        # Agrego los datos al dataframe global del municipio
        df_bga = pd.concat([df_bga, df3], ignore_index=True)

        print("----------------------------------------------------")
        print("El nuevo dataframe está así:")
        df_bga.info()
        print("----------------------------------------------------")

        # Vacío los dataframes para la próxima iteración
        df.drop(df.index, inplace=True)
        df2.drop(df2.index, inplace=True)
        df3.drop(df3.index, inplace=True)

# En este punto ya tengo el dataframe consolidado en la variable df_bga

# Presento en pantalla el resumen de la operación
# Obtener la tabla de frecuencias de la columna 256
tabla_frecuencias = df_bga[256].value_counts()

# Verificar la tabla de frecuencias
print("************************************************")
print("Datos cargados.  Se cargaron datos así:")
print(tabla_frecuencias)
print("************************************************")

# Copio el dataframe eliminando la columna final para cargar los datos al archivo consolidado
df_bga2 = df_bga.iloc[:, 1:-1]


# ******************************************
# *** Escribir en el archivo de destino***
# ******************************************

# Ruta del archivo Excel de destino
archivo_excel = "consolidado.xlsx"

# Abro el libro del archivo Excel
libro_trabajo = openpyxl.load_workbook(archivo_excel)

# Selecciono la hoja del registro
hoja = libro_trabajo["Registro Diario"]

# Leo el contenido de la columna A en la hoja "Registro Diario" antes de escribir nuevos datos
columna_a = [celda.value for celda in hoja["A"]]
# Elimino el primer elemento de la lista para conservar la estructura de la columna A cuando la escriba
columna_a = columna_a[1:]


# Escribir los datos del DataFrame en la hoja "Registro Diario" a partir de la celda B3
for i, fila in enumerate(pd.DataFrame(df_bga2).values):
    for j, valor in enumerate(fila):
        hoja.cell(row=3 + i, column=j + 2, value=valor)  # Escribir los datos a partir de la celda B3

# Conservar el contenido de la columna A en la hoja "Registro Diario"
for i, valor in enumerate(columna_a):
    hoja.cell(row=2 + i, column=1, value=valor)

# Guardar los cambios en el archivo Excel
libro_trabajo.save(archivo_excel)

# Cerrar el archivo
libro_trabajo.close()

print("Terminado. Los datos se guardaron en el archivo consolidado.xlsx")
